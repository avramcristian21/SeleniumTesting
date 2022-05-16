"""
Proiect: GUI Testing
Executarea automata a urmatoarelor teste pe pagina BNR pentru:
    Test 1: verificarea prezentei textului "Banca Naţională a României”
    Test 2: selectarea din drop-down list-ul “Politica monetara" a meniului "Instrumente” si pentru verificarea ajungerii
            pe pagina “Instrumente de politica monetara”
    Test 3: selectarea si verificarea selectiei evenimentului din 24 Mai de pe pagina "Calendarul comunicatelor de presa"
Autor: Avram Cristian-Calin
"""
# Module utilizate
import sys
import time
import variables as v
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service


# Momentul inceperii executiei
start_time = time.time()

# Informatii eroare
def info_error():
    """
    Functie care afiseaza tipul de eroare si linia unde apare
    :return: None
    """
    error = sys.exc_info()[0]
    tb = sys.exc_info()[2]
    print("{} line {}".format(error, tb.tb_lineno))

# Vizualizare excel
def excel_vis():
    """
    Aceasta functie afiseaza informatii despre fisierul excel
    :return: None
    """
    contor = 1
    i = 0
    var = True
    try:
        excel_workbook = load_workbook(v.excel_file)
    except:
        var = False
        info_error()
    if var:
        excel_sheet = excel_workbook.active
        row = excel_sheet[v.line]
        print("Sheets: {}".format(excel_workbook.sheetnames))
        for k in row:
            if k.value != None:
                print("Column: {}".format(v.index[i]))
                col = excel_sheet[v.index[i]]
                for x in col:
                    if x.value != None:
                        print("{}{}: {}".format(v.index[i], contor, x.value))
                        contor = contor + 1
                contor = 1
                i = i + 1

# Citirea din excel
def excel_reading(cell):
    """
    Functie care permite citirea datelor din fisierul excel
    :param cell: 'B2', 'C2', 'D2', type: str
    :return: None
    """
    try:
        excel_workbook = load_workbook(v.excel_file)
        excel_sheet = excel_workbook.active
        return excel_sheet[cell].value
    except:
        info_error()

# Scrierea in excel
def excel_writing(result, cell):
    """
    Functie care permite introducerea unor date in diferite celule din excel
    :param result: Passed or Failed, type: str
    :param cell: 'E2', 'F2', 'G2', type: str
    :return: None
    """
    try:
        excel_workbook = load_workbook(v.excel_file)
        excel_sheet = excel_workbook.active
        excel_sheet[cell] = result
        excel_workbook.save(v.excel_file)
    except:
        info_error()

# Accesarea paginii web
def open_web_page():
    """
    Functie care, odata cu crearea instantei Chrome WebDriver, permite navigarea la pagina BNR prin introducerea
    adresei URL
    :return: driver, type: class
    """
    try:
        s = Service(v.driver_file)
        driver = webdriver.Chrome(service=s)
        driver.maximize_window()
        driver.get(v.url_bnr)
        return driver # driver este variabila de referinta pentru interfata WebDriver
    except:
        info_error()

# Testarea existentei unui text
def first_test(xpath_val, f_cell, s_cell, driver):
    """
    Aceasta functie testeaza daca textul "Banca Naţională a României” se regaseste pe pagina principala
    :param xpath_val: type: str
    :param f_cell: 'B2', type: str
    :param s_cell: 'E2', type: str
    :param driver: type: class
    :return:
    """
    try:
        title = WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, xpath_val))).text
        text_test = excel_reading(f_cell)
        if title == text_test:
            excel_writing(v.result_passed, s_cell)
            print("Passed!")
        else:
            excel_writing(v.result_failed, s_cell)
            print("Failed!")
    except:
        info_error()

    # text_test = excel_reading(f_cell)
    # if title == text_test:
    #     excel_writing(v.result_passed, s_cell)
    #     print("Passed!")
    # else:
    #     excel_writing(v.result_failed, s_cell)
    #     print("Failed!")

# Testarea ajungerii pe o pagina diferita
def second_test(driver):
    """
    Functia data selecteaza un meniu din drop-down-ul "Politica monetara" si verifica daca am ajuns pe pagina respectiva
    :param driver: type: class
    :return: None
    """
    try:
        dropdown = WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_dropdown)))
        achains = ActionChains(driver)
        achains.move_to_element(dropdown).perform()
        WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_element_dd))).click()
        first_test(v.xpath_element_dd_title, 'C2', 'F2', driver)
    except:
        info_error()

# Verificarea unui eveniment
def third_test(driver):
    """
    Aceasta functie are rolul de a selecta, din pagina "Calendarul comunicatelor de presa", un eveniment si verifica
    selectia facuta
    :param driver: type: class
    :return: None
    """
    try:
        WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_bnr))).click()
        WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_element_menu))).click()
        WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_calendar))).click()
        WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_data))).click()
        first_test(v.xpath_data_eveniment, 'D2', 'G2', driver)
    except:
        info_error()

# Alegerea testelor pentru executare
def call():
    """
    Functia call() permite alegerea numarului de teste care vor fi executate
    :return: None
    """
    print("Doriti informatii despre continutul fisierului excel? DA/NU: ")
    info_excel = str(input())
    if info_excel.upper() == "DA":
        print("Continutul fisierului excel:")
        excel_vis()
    else:
        print("Fara informatii!")

    print("-------------------------------------------------------------------------")

    print("Optiuni 1:\n1.Rularea unui singur test\n2.Rularea a doua teste\n3.Rularea celor 3 teste\nAlegeti optiunea:")
    nr_teste = int(input())
    if nr_teste < 4 and nr_teste > 0:
        if nr_teste == 1:
            print("-------------------------------------------------------------------------")
            print("Optiuni 2:\n1.Test 1\n2.Test 2\n3.Test 3\nAlegeti optiunea:")
            option = int(input())
            if option < 4 and option > 0:
                if option == 1:
                    driver = open_web_page()
                    time.sleep(2)
                    first_test(v.xpath_home, 'B2', 'E2', driver)
                    time.sleep(3)
                elif option == 2:
                    driver = open_web_page()
                    time.sleep(2)
                    second_test(driver)
                    time.sleep(3)
                else:
                    driver = open_web_page()
                    time.sleep(2)
                    third_test(driver)
                    time.sleep(3)
            else:
                print("Ati introdus o valoare gresita!")
        elif nr_teste == 2:
            driver = open_web_page()
            time.sleep(2)
            first_test(v.xpath_home, 'B2', 'E2', driver)
            time.sleep(2)
            second_test(driver)
            time.sleep(3)
        else:
            driver = open_web_page()
            time.sleep(2)
            first_test(v.xpath_home, 'B2', 'E2', driver)
            time.sleep(2)
            second_test(driver)
            time.sleep(2)
            WebDriverWait(driver, timeout=5).until(ec.presence_of_element_located((By.XPATH, v.xpath_home))).click()
            time.sleep(2)
            third_test(driver)
            time.sleep(3)
    else:
        print("Ati introdus o valoare gresita!")


call()
execution_time = time.time() - start_time
print("Total execution time {}s".format(execution_time))
